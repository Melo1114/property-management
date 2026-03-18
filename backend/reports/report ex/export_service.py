"""
reports/export_service.py
──────────────────────────
Generates PDF (via reportlab) and Excel (via openpyxl) exports
for all four report types: Dashboard, Rent Roll, Occupancy, Overdue.

Usage in views:
    from .export_service import export_pdf, export_excel
    return export_pdf("rent_roll", data, filename="rent_roll.pdf")
    return export_excel("rent_roll", data, filename="rent_roll.xlsx")
"""

import io
from datetime import date
from decimal import Decimal

from django.http import HttpResponse

# ── ReportLab ─────────────────────────────────────────────────────────────────
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    HRFlowable,
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# ── openpyxl ──────────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter


# ── Brand colours ─────────────────────────────────────────────────────────────
BRAND_DARK   = colors.HexColor("#1E3A5F")   # navy
BRAND_MID    = colors.HexColor("#2E75B6")   # blue
BRAND_LIGHT  = colors.HexColor("#D9E8F5")   # pale blue
BRAND_ACCENT = colors.HexColor("#E8534A")   # red (for overdue/alerts)
WHITE        = colors.white
GREY_ROW     = colors.HexColor("#F5F7FA")

XL_HEADER_FILL = "1E3A5F"
XL_SUBHEADER   = "2E75B6"
XL_ALT_ROW     = "F5F7FA"
XL_ACCENT      = "E8534A"


# ═══════════════════════════════════════════════════════════════════════════════
#  SHARED HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _pdf_response(filename: str) -> tuple[io.BytesIO, HttpResponse]:
    buffer = io.BytesIO()
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return buffer, response


def _excel_response(filename: str) -> tuple[Workbook, HttpResponse]:
    wb = Workbook()
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return wb, response


def _pdf_header_elements(title: str, subtitle: str = "") -> list:
    """Returns a list of flowables: logo placeholder, title, subtitle, divider."""
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Normal"],
        fontSize=18,
        fontName="Helvetica-Bold",
        textColor=BRAND_DARK,
        spaceAfter=4,
    )
    sub_style = ParagraphStyle(
        "ReportSub",
        parent=styles["Normal"],
        fontSize=10,
        fontName="Helvetica",
        textColor=colors.HexColor("#555555"),
        spaceAfter=8,
    )
    elements = [
        Paragraph("PROPERTY MANAGEMENT SYSTEM", ParagraphStyle(
            "Brand", parent=styles["Normal"],
            fontSize=9, fontName="Helvetica-Bold",
            textColor=BRAND_MID, spaceAfter=2,
        )),
        Paragraph(title, title_style),
    ]
    if subtitle:
        elements.append(Paragraph(subtitle, sub_style))
    elements.append(HRFlowable(width="100%", thickness=2, color=BRAND_MID, spaceAfter=10))
    return elements


def _pdf_table_style(has_totals_row: bool = False) -> TableStyle:
    style = [
        # Header row
        ("BACKGROUND",    (0, 0), (-1, 0),  BRAND_DARK),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  WHITE),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0),  9),
        ("ALIGN",         (0, 0), (-1, 0),  "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, 0),  8),
        ("TOPPADDING",    (0, 0), (-1, 0),  8),
        # Data rows
        ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 1), (-1, -1), 8),
        ("TOPPADDING",    (0, 1), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 5),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, GREY_ROW]),
        # Grid
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("LINEBELOW",     (0, 0), (-1, 0),  1.5, BRAND_MID),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]
    if has_totals_row:
        style += [
            ("BACKGROUND",  (0, -1), (-1, -1), BRAND_LIGHT),
            ("FONTNAME",    (0, -1), (-1, -1), "Helvetica-Bold"),
            ("LINEABOVE",   (0, -1), (-1, -1), 1.5, BRAND_MID),
        ]
    return TableStyle(style)


def _xl_header_row(ws, row: int, headers: list, col_start: int = 1):
    """Write a styled header row to an Excel worksheet."""
    fill   = PatternFill("solid", fgColor=XL_HEADER_FILL)
    font   = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        bottom=Side(style="medium", color="2E75B6"),
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
    )
    for i, header in enumerate(headers, start=col_start):
        cell = ws.cell(row=row, column=i, value=header)
        cell.font   = font
        cell.fill   = fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _xl_title_block(ws, title: str, subtitle: str, num_cols: int):
    """Write report title block in first two rows, merge across all columns."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font      = Font(bold=True, size=14, color=XL_HEADER_FILL)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    sub_cell = ws.cell(row=2, column=1, value=subtitle)
    sub_cell.font      = Font(italic=True, size=9, color="555555")
    sub_cell.alignment = Alignment(horizontal="center")

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 16


def _xl_data_row(ws, row: int, values: list, alt: bool = False):
    """Write a data row with optional alternating background."""
    fill   = PatternFill("solid", fgColor=XL_ALT_ROW) if alt else None
    border = Border(
        left=Side(style="thin",   color="CCCCCC"),
        right=Side(style="thin",  color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    for i, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=i, value=value)
        cell.border = border
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(vertical="center")


def _xl_auto_width(ws, min_width: int = 12, max_width: int = 40):
    """Auto-size column widths based on content."""
    for col_cells in ws.columns:
        length = max(
            (len(str(cell.value)) if cell.value else 0)
            for cell in col_cells
        )
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 2, min_width), max_width)


# ═══════════════════════════════════════════════════════════════════════════════
#  DASHBOARD REPORT
# ═══════════════════════════════════════════════════════════════════════════════

def dashboard_pdf(data: dict) -> HttpResponse:
    buffer, response = _pdf_response("dashboard_report.pdf")
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    today    = date.today().strftime("%d %B %Y")
    elements = _pdf_header_elements("Dashboard Report", f"Generated: {today}")
    elements.append(Spacer(1, 0.5*cm))

    summary_data = [
        ["Metric", "Value"],
        ["Total Properties",           str(data.get("total_properties", 0))],
        ["Total Units",                str(data.get("total_units", 0))],
        ["Occupied Units",             str(data.get("occupied_units", 0))],
        ["Vacant Units",               str(data.get("vacant_units", 0))],
        ["Units Under Maintenance",    str(data.get("maintenance_units", 0))],
        ["Open Maintenance Requests",  str(data.get("open_maintenance_requests", 0))],
        ["Overdue Invoices (count)",   str(data.get("overdue_invoices_count", 0))],
        ["Overdue Invoices (total)",   f"R {data.get('overdue_invoices_total', '0.00')}"],
    ]

    t = Table(summary_data, colWidths=[10*cm, 7*cm])
    t.setStyle(_pdf_table_style())
    elements.append(t)

    doc.build(elements)
    response.write(buffer.getvalue())
    return response


def dashboard_excel(data: dict) -> HttpResponse:
    wb, response = _excel_response("dashboard_report.xlsx")
    ws = wb.active
    ws.title = "Dashboard"
    today = date.today().strftime("%d %B %Y")
    _xl_title_block(ws, "Dashboard Report", f"Generated: {today}", 2)

    headers = ["Metric", "Value"]
    _xl_header_row(ws, 3, headers)

    rows = [
        ("Total Properties",          data.get("total_properties", 0)),
        ("Total Units",               data.get("total_units", 0)),
        ("Occupied Units",            data.get("occupied_units", 0)),
        ("Vacant Units",              data.get("vacant_units", 0)),
        ("Units Under Maintenance",   data.get("maintenance_units", 0)),
        ("Open Maintenance Requests", data.get("open_maintenance_requests", 0)),
        ("Overdue Invoices (count)",  data.get("overdue_invoices_count", 0)),
        ("Overdue Invoices (total)",  f"R {data.get('overdue_invoices_total', '0.00')}"),
    ]
    for i, (metric, value) in enumerate(rows, start=4):
        _xl_data_row(ws, i, [metric, value], alt=(i % 2 == 0))

    _xl_auto_width(ws)
    wb.save(response)
    return response


# ═══════════════════════════════════════════════════════════════════════════════
#  RENT ROLL REPORT
# ═══════════════════════════════════════════════════════════════════════════════

def rent_roll_pdf(data: dict) -> HttpResponse:
    buffer, response = _pdf_response("rent_roll.pdf")
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    today    = date.today().strftime("%d %B %Y")
    elements = _pdf_header_elements("Rent Roll Report", f"Generated: {today}")

    rows = data.get("rent_roll", [])
    total_rent = sum(Decimal(r.get("monthly_rent", "0") or "0") for r in rows)

    table_data = [[
        "Property", "Unit", "Status",
        "Monthly Rent", "Tenant", "Lease End",
    ]]
    for r in rows:
        table_data.append([
            r.get("property_name", ""),
            r.get("unit_number", ""),
            r.get("status", ""),
            f"R {r.get('monthly_rent', '0.00')}",
            r.get("tenant_name") or "—",
            r.get("lease_end_date") or "—",
        ])
    # Totals row
    table_data.append([
        "TOTAL", f"{len(rows)} units", "",
        f"R {total_rent:,.2f}", "", "",
    ])

    col_widths = [6*cm, 2.5*cm, 3*cm, 3.5*cm, 5*cm, 3*cm]
    t = Table(table_data, colWidths=col_widths)
    t.setStyle(_pdf_table_style(has_totals_row=True))
    elements.append(t)

    doc.build(elements)
    response.write(buffer.getvalue())
    return response


def rent_roll_excel(data: dict) -> HttpResponse:
    wb, response = _excel_response("rent_roll.xlsx")
    ws = wb.active
    ws.title = "Rent Roll"
    today = date.today().strftime("%d %B %Y")

    headers = ["Property", "Unit", "Status", "Monthly Rent (R)", "Tenant", "Lease End Date"]
    _xl_title_block(ws, "Rent Roll Report", f"Generated: {today}", len(headers))
    _xl_header_row(ws, 3, headers)

    rows = data.get("rent_roll", [])
    total_rent = Decimal("0")

    for i, r in enumerate(rows, start=4):
        rent = Decimal(r.get("monthly_rent", "0") or "0")
        total_rent += rent
        _xl_data_row(ws, i, [
            r.get("property_name", ""),
            r.get("unit_number", ""),
            r.get("status", ""),
            float(rent),
            r.get("tenant_name") or "—",
            r.get("lease_end_date") or "—",
        ], alt=(i % 2 == 0))
        # Format rent column as currency
        ws.cell(row=i, column=4).number_format = 'R#,##0.00'

    # Totals row
    totals_row = len(rows) + 4
    _xl_data_row(ws, totals_row, ["TOTAL", f"{len(rows)} units", "", float(total_rent), "", ""])
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=totals_row, column=col)
        cell.font = Font(bold=True, color=XL_HEADER_FILL)
        cell.fill = PatternFill("solid", fgColor="D9E8F5")
    ws.cell(row=totals_row, column=4).number_format = 'R#,##0.00'

    _xl_auto_width(ws)
    wb.save(response)
    return response


# ═══════════════════════════════════════════════════════════════════════════════
#  OCCUPANCY REPORT
# ═══════════════════════════════════════════════════════════════════════════════

def occupancy_pdf(data: dict) -> HttpResponse:
    buffer, response = _pdf_response("occupancy_report.pdf")
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    today    = date.today().strftime("%d %B %Y")
    elements = _pdf_header_elements("Occupancy Report", f"Generated: {today}")
    elements.append(Spacer(1, 0.5*cm))

    summary = [
        ["Metric",                "Units",   "Percentage"],
        ["Occupied",
         str(data.get("occupied_units", 0)),
         f"{data.get('occupancy_rate_percent', 0)}%"],
        ["Vacant",
         str(data.get("vacant_units", 0)),
         f"{round((data.get('vacant_units', 0) / data.get('total_units', 1)) * 100, 2)}%"
         if data.get("total_units") else "0%"],
        ["Under Maintenance",
         str(data.get("maintenance_units", 0)),
         f"{round((data.get('maintenance_units', 0) / data.get('total_units', 1)) * 100, 2)}%"
         if data.get("total_units") else "0%"],
        ["TOTAL",
         str(data.get("total_units", 0)), "100%"],
    ]
    t = Table(summary, colWidths=[8*cm, 4*cm, 4*cm])
    t.setStyle(_pdf_table_style(has_totals_row=True))
    elements.append(t)

    doc.build(elements)
    response.write(buffer.getvalue())
    return response


def occupancy_excel(data: dict) -> HttpResponse:
    wb, response = _excel_response("occupancy_report.xlsx")
    ws = wb.active
    ws.title = "Occupancy"
    today = date.today().strftime("%d %B %Y")
    total = data.get("total_units", 1) or 1

    _xl_title_block(ws, "Occupancy Report", f"Generated: {today}", 3)
    _xl_header_row(ws, 3, ["Metric", "Units", "Percentage"])

    rows = [
        ("Occupied",          data.get("occupied_units", 0),    data.get("occupancy_rate_percent", 0) / 100),
        ("Vacant",            data.get("vacant_units", 0),      data.get("vacant_units", 0) / total),
        ("Under Maintenance", data.get("maintenance_units", 0), data.get("maintenance_units", 0) / total),
        ("TOTAL",             data.get("total_units", 0),       1.0),
    ]
    for i, (label, units, pct) in enumerate(rows, start=4):
        _xl_data_row(ws, i, [label, units, pct], alt=(i % 2 == 0))
        ws.cell(row=i, column=3).number_format = '0.00%'
        if label == "TOTAL":
            for col in range(1, 4):
                ws.cell(row=i, column=col).font = Font(bold=True, color=XL_HEADER_FILL)
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor="D9E8F5")

    _xl_auto_width(ws)
    wb.save(response)
    return response


# ═══════════════════════════════════════════════════════════════════════════════
#  OVERDUE REPORT
# ═══════════════════════════════════════════════════════════════════════════════

def overdue_pdf(data: dict) -> HttpResponse:
    buffer, response = _pdf_response("overdue_report.pdf")
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    today    = date.today().strftime("%d %B %Y")
    elements = _pdf_header_elements(
        "Overdue Invoices Report",
        f"Generated: {today}  |  Total Overdue: R {data.get('total_overdue', '0.00')}  "
        f"|  Count: {data.get('count', 0)}"
    )

    invoices = data.get("overdue_invoices", [])
    table_data = [[
        "Invoice #", "Property", "Unit",
        "Tenant", "Due Date", "Invoice Amount", "Balance Owed",
    ]]
    for inv in invoices:
        table_data.append([
            str(inv.get("invoice_id", "")),
            inv.get("property_name", ""),
            inv.get("unit_number", ""),
            inv.get("tenant_name") or inv.get("tenant_username") or "—",
            str(inv.get("due_date", "")),
            f"R {inv.get('amount', '0.00')}",
            f"R {inv.get('balance', '0.00')}",
        ])

    total_overdue = data.get("total_overdue", "0.00")
    table_data.append([
        "", "", "", f"{len(invoices)} invoices", "",
        "TOTAL OVERDUE:", f"R {total_overdue}",
    ])

    col_widths = [2.5*cm, 6*cm, 2.5*cm, 5*cm, 3*cm, 4*cm, 4*cm]
    t = Table(table_data, colWidths=col_widths)

    # Apply red highlight to balance column
    style = _pdf_table_style(has_totals_row=True)
    style.add("TEXTCOLOR", (6, 1), (6, -2), BRAND_ACCENT)
    style.add("FONTNAME",  (6, 1), (6, -2), "Helvetica-Bold")
    t.setStyle(style)

    elements.append(t)
    doc.build(elements)
    response.write(buffer.getvalue())
    return response


def overdue_excel(data: dict) -> HttpResponse:
    wb, response = _excel_response("overdue_report.xlsx")
    ws = wb.active
    ws.title = "Overdue Invoices"
    today = date.today().strftime("%d %B %Y")

    headers = [
        "Invoice #", "Property", "Unit",
        "Tenant", "Due Date", "Invoice Amount (R)", "Balance Owed (R)",
    ]
    _xl_title_block(
        ws,
        "Overdue Invoices Report",
        f"Generated: {today}  |  Total Overdue: R {data.get('total_overdue', '0.00')}",
        len(headers),
    )
    _xl_header_row(ws, 3, headers)

    invoices = data.get("overdue_invoices", [])
    red_font   = Font(bold=True, color=XL_ACCENT)
    red_fill   = PatternFill("solid", fgColor="FDE8E7")

    for i, inv in enumerate(invoices, start=4):
        _xl_data_row(ws, i, [
            inv.get("invoice_id", ""),
            inv.get("property_name", ""),
            inv.get("unit_number", ""),
            inv.get("tenant_name") or inv.get("tenant_username") or "—",
            inv.get("due_date", ""),
            float(Decimal(inv.get("amount",  "0") or "0")),
            float(Decimal(inv.get("balance", "0") or "0")),
        ], alt=(i % 2 == 0))
        # Highlight balance column in red
        balance_cell = ws.cell(row=i, column=7)
        balance_cell.font   = red_font
        balance_cell.fill   = red_fill
        balance_cell.number_format = 'R#,##0.00'
        ws.cell(row=i, column=6).number_format = 'R#,##0.00'

    # Totals row
    totals_row = len(invoices) + 4
    _xl_data_row(ws, totals_row, [
        "", "", "", f"{len(invoices)} invoices", "",
        "TOTAL OVERDUE:",
        float(Decimal(str(data.get("total_overdue", "0")))),
    ])
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=totals_row, column=col)
        cell.font = Font(bold=True, color=XL_HEADER_FILL)
        cell.fill = PatternFill("solid", fgColor="D9E8F5")
    ws.cell(row=totals_row, column=7).number_format = 'R#,##0.00'

    _xl_auto_width(ws)
    wb.save(response)
    return response


# ═══════════════════════════════════════════════════════════════════════════════
#  PUBLIC DISPATCH FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════

PDF_RENDERERS = {
    "dashboard": dashboard_pdf,
    "rent_roll":  rent_roll_pdf,
    "occupancy":  occupancy_pdf,
    "overdue":    overdue_pdf,
}

EXCEL_RENDERERS = {
    "dashboard": dashboard_excel,
    "rent_roll":  rent_roll_excel,
    "occupancy":  occupancy_excel,
    "overdue":    overdue_excel,
}


def export_pdf(report_type: str, data: dict) -> HttpResponse:
    renderer = PDF_RENDERERS.get(report_type)
    if not renderer:
        raise ValueError(f"Unknown report type: {report_type}")
    return renderer(data)


def export_excel(report_type: str, data: dict) -> HttpResponse:
    renderer = EXCEL_RENDERERS.get(report_type)
    if not renderer:
        raise ValueError(f"Unknown report type: {report_type}")
    return renderer(data)
